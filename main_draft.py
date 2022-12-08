# Decoding the organization's IP address from excel using ipinfo
import urllib3
import random
from requests import get
from pathlib import Path
import openpyxl
from bs4 import BeautifulSoup


def save_ip_list(ls, file_name):
    try:
        with open(file_name, 'w') as w_file:
            for ip in ls:
                w_file.write(ip + '\n')
    except Exception as e:
        print(f'Error writing the IP list to the file: {file_name} ({str(e)})')


def get_ip_list_from_file(file_name):
    try:
        xlsx_file = Path(file_name)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        ip_set = set()
        # IPs are located in 11 column of example.xlsx
        # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=12, max_col=12):
        for row in sheet.iter_rows(min_row=2, max_row=20000, min_col=12, max_col=12):
            for cell in row:
                if cell.value is not None:
                    ip_set.add(cell.value)
                    # print(cell.value, end=" ")
                    # print()
        print(ip_set)
        return list(ip_set)
    except Exception as e:
        print(f'Error handling file: {file_name} ({str(e)})')


def execution_requests(url, ip_ls_file, res_file):
    with open(ip_ls_file, 'r') as r_file:
        with open(res_file, 'w') as w_file_200:
            with open('unable_to_resolve.txt', 'w') as ex_file_429:
                cnt_200 = 0
                cnt_429 = 0
                un_ls = list()
                for ip in r_file.readlines():
                    ip = ip.strip()
                    try:
                        print('Resolving : ' + ip)
                        # user_agent_list = [
                        #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.1 Safari/605.1.15',
                        #     'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0',
                        #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36',
                        #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:77.0) Gecko/20100101 Firefox/77.0',
                        #     'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36',
                        # ]
                        user_agent_list = [
                            'Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)',
                            'Opera/9.80 (X11; Linux i686; U; ru) Presto/2.8.131 Version/11.11',
                            'Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:15.0) Gecko/20100101 Firefox/15.0.1',
                            'Mozilla/5.0 (iPad; CPU OS 6_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/6.0 Mobile/10A5355d Safari/8536.25',
                        ]
                        for i in range(1, 4):
                            user_agent = random.choice(user_agent_list)
                        headers = {'User-Agent': user_agent}
                        # headers = {'User-Agent':
                        #                'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}
                        # headers = {'User-Agent':
                        #                'Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 83.0.4103.97 Safari / 537.36',
                        #            'content-type': 'application/json; charset=utf-8',
                        #            'Content-type': 'application/json; charset=utf-8',
                        #            'X-Content-Type-Options': 'nosniff'}
                        # headers = {'content-type': 'application/json; charset=utf-8'}
                        # headers = {'User-Agent': 'curl/7.65.3'}

                        # proxies = {
                        #     "http": "http://token:render_js=False&premium_proxy=True@proxy.scrapingbee.com:8886",
                        #     "https": "https://token:render_js=False&premium_proxy=True@proxy.scrapingbee.com:8887"
                        # }
                        proxies = {
                            "https": "24.37.245.42:51056"
                        }
                        urllib3.disable_warnings()
                        response = get(url + ip, proxies=proxies, verify=False)
                        # response = get(url + ip, headers=headers)
                        # response = get(url + ip)
                        # sleep(3)
                        print(response.status_code)
                        # work with JSON
                        if response.status_code == 200:
                            w_file_200.write(f"{ip}: "
                                             f"{response.json()['city'] if 'city' in response.json() else 'No City'}, "
                                             f"{response.json()['country'] if 'country' in response.json() else 'No Country'}, "
                                             f"{response.json()['org'] if 'org' in response.json() else 'No Organisation'}, "
                                             f"{response.json()['loc'] if 'loc' in response.json() else 'No Location'}, "
                                             f"{response.json()['hostname'] if 'hostname' in response.json() else 'No Hostname'}"
                                             f"\n")
                            cnt_200 += 1
                        # work with HTML
                        if response.status_code == 429:
                            soup = BeautifulSoup(response.text, 'html.parser')
                            geo_table = str(soup.find('table', class_='table table-borderless table-xs geo-table'))
                            table_data = dict([[cell.text for cell in row('td')]
                                               for row in BeautifulSoup(geo_table, 'html.parser')('tr')])
                            print(table_data)
                            ex_file_429.write(f"{ip}: "
                                              f"{table_data['City'] if 'City' in table_data else 'No City'}, "
                                              f"{table_data['Country'] if 'Country' in table_data else 'No Country'}, "
                                              f"{table_data['org'] if 'org' in table_data else 'No Organisation'}, "
                                              f"{table_data['Coordinates'] if 'Coordinates' in table_data else 'No Location'}, "
                                              f"{table_data['hostname'] if 'hostname' in table_data else 'No Hostname'}"
                                              f"\n")
                            cnt_429 += 1
                    except Exception as e:
                        print(f'Request execution error: {url + ip} ({str(e)})')
                        # f'\n{response.text}')
                    #     un_ls.append(ip)
                    # for ip in un_ls:
                    #     ex_file.write(f"{ip}: "
                    #                   f"\n")
                print(f'Quantity resolved: {cnt_200}')
                print(f'Quantity unable to resolve: {cnt_429}')


if __name__ == '__main__':
    input_file = 'example.xlsx'
    ip_file = 'ip.txt'
    result_file = 'result.txt'
    ipinfo = 'https://ipinfo.io/'
    # ip_list = get_ip_list_from_file(input_file)
    # save_ip_list(ip_list, ip_file)
    execution_requests(ipinfo, ip_file, result_file)
