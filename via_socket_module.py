# Decoding IP address to domain name from excel using socket module
import socket
import openpyxl
from pathlib import Path


def use_socket():
    xlsx_file = Path('example.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    ip_set = set()
    # IPs are located in 11 column of example.xlsx
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=12, max_col=12):
    # for row in sheet.iter_rows(min_row=2, max_row=10000, min_col=12, max_col=12):
        for cell in row:
            if cell.value is not None:
                ip_set.add(cell.value)
                # print(cell.value, end=" ")
                # print()
    print(ip_set)
    with open(r'ip_address.txt', 'w') as w_file:
        for ip in ip_set:
            w_file.write(ip + '\n')

    with open(r'ip_address.txt', 'r') as r_file:
        with open(r'domain_name.txt', 'w') as w_file:
            un_ls = list()
            for ip in r_file.readlines():
                ip = ip.strip()
                try:
                    print('Resolving : ' + ip)
                    for i in range(1):
                        hostname = socket.gethostbyaddr(ip)
                        print(hostname)
                        w_file.write(ip + ' - ' + hostname[0] + '\n')
                except:
                    print('Unable to resolve: ' + ip)
                    un_ls.append(ip)
            with open(r'unable_to_resolve.txt', 'w') as ex_file:
                for ip in un_ls:
                    ex_file.write('Unable to resolve: ' + ip + '\n')
